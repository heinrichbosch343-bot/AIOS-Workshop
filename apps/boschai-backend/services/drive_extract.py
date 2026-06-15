"""
Extract readable text from a Google Drive file.

Handles the file types Heinrich actually stores: Google Docs/Sheets (exported),
PDFs (pypdf), Word .docx (python-docx), Excel .xlsx (openpyxl), and plain text.
Returns plain text so the dashboard can preview it and the brain can use it.
"""
import io

from googleapiclient.discovery import build

from services.drive import get_credentials, download_file

# Google-native types must be EXPORTED, not downloaded directly.
_GOOGLE_EXPORT = {
    "application/vnd.google-apps.document": "text/plain",
    "application/vnd.google-apps.spreadsheet": "text/csv",
    "application/vnd.google-apps.presentation": "text/plain",
}

_PDF = "application/pdf"
_DOCX = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MAX_CHARS = 20000  # cap so a huge file can't blow up the response / a prompt


def _export_google(file_id: str, export_mime: str) -> str:
    service = build("drive", "v3", credentials=get_credentials())
    data = service.files().export_media(fileId=file_id, mimeType=export_mime).execute()
    return data.decode("utf-8", "ignore") if isinstance(data, bytes) else str(data)


def _pdf_to_text(raw: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(raw))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _docx_to_text(raw: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(raw))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _xlsx_to_text(raw: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(", ".join(cells))
    return "\n".join(lines)


def extract_text(file_id: str, mime_type: str = "", name: str = "",
                 max_chars: "int | None" = MAX_CHARS) -> dict:
    """Return {text, chars, truncated, mime}. Never raises — errors come back as text.

    max_chars caps the returned text (default MAX_CHARS for previews/queries). Pass None to
    get the full text — the indexer needs whole documents, not a 20K preview.
    """
    try:
        if mime_type in _GOOGLE_EXPORT:
            text = _export_google(file_id, _GOOGLE_EXPORT[mime_type])
        elif mime_type == _PDF or name.lower().endswith(".pdf"):
            text = _pdf_to_text(download_file(file_id))
        elif mime_type == _DOCX or name.lower().endswith(".docx"):
            text = _docx_to_text(download_file(file_id))
        elif mime_type == _XLSX or name.lower().endswith(".xlsx"):
            text = _xlsx_to_text(download_file(file_id))
        elif mime_type.startswith("text/") or name.lower().endswith((".txt", ".md", ".csv")):
            text = download_file(file_id).decode("utf-8", "ignore")
        else:
            return {"text": f"(Preview not supported for this file type: {mime_type or name})",
                    "chars": 0, "truncated": False, "mime": mime_type}
    except Exception as e:
        return {"text": f"(Could not extract content: {e})", "chars": 0, "truncated": False, "mime": mime_type}

    text = (text or "").strip()
    full = len(text)
    cap = full if max_chars is None else max_chars
    truncated = full > cap
    return {"text": text[:cap], "chars": full, "truncated": truncated, "mime": mime_type}
